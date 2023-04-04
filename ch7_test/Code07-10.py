import openpyxl
from copy import copy

workbook = openpyxl.load_workbook('Excel/singer.xlsx')
wsheetList = workbook.sheetnames # 엑셀 파일 시트 이름 리스트로

outWorkbook = openpyxl.Workbook() # 기본으로 빈 엑셀 파일(워크북) 생성
outWorkbook.remove(outWorkbook['Sheet']) # 기본으로 생성된 시트를 일단 제거

for wsName in wsheetList : # 엑셀 파일 시트 이름이 리스트로 된 시트명을 하나씩 꺼내서 담기.
    worksheet = workbook[wsName] # 해당 시트명으로 워크북에서 선택.
    outSheet = outWorkbook.create_sheet(wsName) # 쓰여질 워크북에 해당 시트이름으로 생성
    for row in range(1, worksheet.max_row+1) : # 0행이 아니라 1행부터, 끝부분도 +1 해주기.
        outSheet.row_dimensions[row].height = worksheet.row_dimensions[row].height
        for col in range(1, worksheet.max_column+1) :
            # ord('A') -> 숫자로, chr(65) 다시 문자로 
            outSheet.column_dimensions[chr(ord('A')+col-1)].width \
                = worksheet.column_dimensions[chr(ord('A')+col-1)].width
                # 입력 워크시트의 현재 셀
            inCell = worksheet.cell(row=row, column=col)
            # 출력 워크시트의 셀
            outCell = outSheet.cell(row=row, column=col, value= inCell.value)
            if inCell.has_style:
                outCell.font = copy(inCell.font)
                outCell.border = copy(inCell.border)
                outCell.fill = copy(inCell.fill)
                outCell.number_format = copy(inCell.number_format)
                outCell.alignment = copy(inCell.alignment)

outWorkbook.save('Excel/singer_copy.xlsx')
print("Save. OK~")